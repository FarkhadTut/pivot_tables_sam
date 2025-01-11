from openpyxl.utils import get_column_letter


def getIndexes(df, value):
    # Empty list
    listOfPos = []

    # isin() method will return a dataframe with
    # boolean values, True at the positions   
    # where element exists
    result = df.isin([value])

    # any() method will return
    # a boolean series
    seriesObj = result.any()

    # Get list of column names where
    # element exists
    columnNames = list(seriesObj[seriesObj == True].index)

    # Iterate over the list of columns and
    # extract the row index where element exists
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)

        for row in rows:
            listOfPos.append((row, col))

    # This list contains a list tuples with
    # the index of element in the dataframe
    return listOfPos


def add_sum_formula(df_total, total_word):
    ### summing with Excel formula
    columns = df_total.columns.values.tolist()
    col_total_idx = 1
    total_row_idx = None
    for idx, row in df_total.iterrows():
        if str(row[columns[col_total_idx]]).upper() == total_word:
            if total_row_idx is not None:
                for col in df_total.columns.values.tolist()[2:]:
                    col_idx = columns.index(col) 
                    col_letter = get_column_letter(col_idx+1)
                    sum_formula = f'=sum({col_letter}{total_row_idx+1}:{col_letter}{idx-1})'
                    df_total.at[total_row_idx, col] = sum_formula
            total_row_idx = idx
      
    else:
        mask = df_total[columns[col_total_idx]].str.upper() == total_word
        last_total_row_idx = df_total[mask].index.values[-1] 
        idx_last = df_total.index.values[-1]
        for col in df_total.columns.values.tolist()[2:]:
            col_idx = columns.index(col) 
            col_letter = get_column_letter(col_idx+1)
            sum_formula = f'=sum({col_letter}{last_total_row_idx+1}:{col_letter}{idx_last})'
            df_total.at[last_total_row_idx, col] = sum_formula

    return df_total


def add_sum_formula_table_1(df_total, total_word):
    ### summing with Excel formula
    ksz_eiz_stsz = [
        'ЭИЗлар бўйича',
        'КСЗлар  бўйича',
        'ЁСТЗ бўйича',
        '"Ургут" ЭИЗ участкалари бўйича',
        'КСЗлар бўйича',
    ]
    columns = df_total.columns.values.tolist()
    col_total_idx = 1
    total_row_idx = None
    for idx, row in df_total.iterrows():
        if str(row[columns[col_total_idx]]).upper() == total_word:
            if total_row_idx is not None:
                for col in df_total.columns.values.tolist()[2:]:
                    col_idx = columns.index(col) 
                    col_letter = get_column_letter(col_idx+1)
                    if df_total.at[idx-1, columns[col_total_idx]] in ksz_eiz_stsz:
                        sum_formula = f'=sum({col_letter}{total_row_idx+1}:{col_letter}{idx-2})'
                    else:
                        sum_formula = f'=sum({col_letter}{total_row_idx+1}:{col_letter}{idx-1})'

                    df_total.at[total_row_idx, col] = sum_formula
            total_row_idx = idx
      
    else:
        mask = df_total[columns[col_total_idx]].str.upper() == total_word
        last_total_row_idx = df_total[mask].index.values[-1] 
        idx_last = df_total.index.values[-1]
        for col in df_total.columns.values.tolist()[2:]:
            col_idx = columns.index(col) 
            col_letter = get_column_letter(col_idx+1)
            sum_formula = f'=sum({col_letter}{last_total_row_idx+1}:{col_letter}{idx_last})'
            df_total.at[last_total_row_idx, col] = sum_formula
    return df_total



def add_sum_formula(df_total, total_word, col_total_idx=1, sum_start_col_idx=2):
    ### summing with Excel formula
    columns = df_total.columns.values.tolist()
    total_row_idx = None
    for idx, row in df_total.iterrows():
        if str(row[columns[col_total_idx]]).upper() == total_word:
            if total_row_idx is not None:
                for col in df_total.columns.values.tolist()[sum_start_col_idx:]:
                    col_idx = columns.index(col) 
                    col_letter = get_column_letter(col_idx+1)
                    sum_formula = f'=sum({col_letter}{total_row_idx+3}:{col_letter}{idx+1})'
                    df_total.at[total_row_idx, col] = sum_formula
            total_row_idx = idx
      
    else:
        mask = df_total[columns[col_total_idx]].str.upper() == total_word
        last_total_row_idx = df_total[mask].index.values[-1] 
        idx_last = df_total.index.values[-1]
        for col in df_total.columns.values.tolist()[sum_start_col_idx:]:
            col_idx = columns.index(col) 
            col_letter = get_column_letter(col_idx+1)
            sum_formula = f'=sum({col_letter}{last_total_row_idx+3}:{col_letter}{idx_last+2})'
            df_total.at[last_total_row_idx, col] = sum_formula

    return df_total


def add_ksz_sum_formula_table_1(df_total, total_word):
    
    ### summing with Excel formula
    ksz_eiz_stsz = [
        'ЭИЗлар бўйича',
        'КСЗлар  бўйича',
        'ЁСТЗ бўйича',
        '"Ургут" ЭИЗ участкалари бўйича',
        'КСЗлар бўйича',
    ]


    columns = df_total.columns.values.tolist()
    col_total_idx = 1
    col_total = columns[col_total_idx]
    mask = ((df_total[col_total].isin(ksz_eiz_stsz)) | (df_total[col_total].str.upper() == total_word))
    df_temp = df_total[mask]
    total_row_idx = None
    for idx, row in df_total.iterrows():
        if row[columns[col_total_idx]] in ksz_eiz_stsz:
            if total_row_idx is not None:
                for col in df_total.columns.values.tolist()[2:]:
                    col_idx = columns.index(col) 
                    col_letter = get_column_letter(col_idx+1)
                    mask_indecies = ((df_temp.index >= total_row_idx+1) & (df_temp.index <= idx-2) \
                                    & (df_temp[col_total].str.upper() == total_word)) 
                    indecies = [str(i) for i in df_temp[mask_indecies].index.values.tolist()]
                    indecies_str = col_letter + f' ,{col_letter}'.join(indecies)
                    sum_formula = f'=sum({indecies_str})'

                    df_total.at[total_row_idx, col] = sum_formula
            total_row_idx = idx
      
    else:
        mask = df_total[columns[col_total_idx]].isin(ksz_eiz_stsz)
        last_total_row_idx = df_total[mask].index.values[-1] 
        idx_last = df_total.index.values[-1]
        for col in df_total.columns.values.tolist()[2:]:
            col_idx = columns.index(col) 
            col_letter = get_column_letter(col_idx+1)
            mask_indecies = ((df_temp.index >= total_row_idx+1) & (df_temp.index <= idx-2) \
                            & (df_temp[col_total].str.upper() == total_word)) 
            indecies = [str(i) for i in df_temp[mask_indecies].index.values.tolist()]
            indecies_str = col_letter + f' ,{col_letter}'.join(indecies)
            sum_formula = f'=sum({indecies_str})'


            df_total.at[last_total_row_idx, col] = sum_formula
    return df_total
