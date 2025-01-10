import openpyxl
import os
from glob import glob
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter

merge_walk_horiz = 20
merge_walk_vert = 5

def table_1():
    files = glob(os.path.join(os.path.join('out', '*.xlsx')))
    file = files[0]
    wb = openpyxl.load_workbook(file)
    ws = wb['СЗ-1']
    ws.delete_rows(1)
    ws.delete_rows(1)
    
    ## pure exception just for this table 
    ws.cell(1, column_index_from_string('AB')).value = '.'
    ####
    
    
    for col in range(1, ws.max_column+1):
        for row in range(1, merge_walk_vert):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue 
            cell.alignment = Alignment(wrap_text=True,
                                       horizontal='center',
                                       vertical='center')
            cell.font = Font(bold=True)

######################## MERGE HEADERS #############################################
            horiz_merge = 0
            vert_merge = 0
            for walk_step in range(1, merge_walk_horiz-1): 
                col_next = col+walk_step      
                if col_next > ws.max_column:
                    break
                cell_next = ws.cell(row, col_next)
                if cell_next.value is None:
                    if row != 1:
                        try:
                            for upper_row in range(1, row):
                               
                                if not ws.cell(row-upper_row, col_next).value is None:
                                    raise Exception('break')
                            else:
                                horiz_merge += 1
                        except:
                            break

                           
                    else:
                        horiz_merge += 1
                else:
                    break

            for walk_step in range(1, merge_walk_vert-1): 
                row_next = row+walk_step      
                if row_next > merge_walk_vert-1:
                    break 
                cell_next = ws.cell(row_next, col)
                
                if cell_next.value is None:
                    vert_merge += 1
                else:
                    break
            

            ws.merge_cells(start_row=row,\
                           end_row=row+vert_merge,\
                           start_column=col,
                           end_column=col+horiz_merge)
                
###################################################################################

            
        ws.column_dimensions[get_column_letter(col)].width = 18.7
    
    for row in range(1, merge_walk_vert):   
        ws.row_dimensions[row].height = 37.25


    wb.save(os.path.join('out', 'pretty', os.path.basename(file).replace('.xlsx', '_formatted.xlsx')))
    wb.close()