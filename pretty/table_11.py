import openpyxl
import os
from glob import glob
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
# from .utils import merge_headers
merge_walk_horiz = 20
merge_walk_vert = 2

FILES_STARTWITH = '11.'

def table_11():
    files = glob(os.path.join(os.path.join('out', '*.xlsx')))
    file = [f for f in files if os.path.basename(f).startswith(FILES_STARTWITH)][0]
    filename_out = os.path.join('out', 'pretty', os.path.basename(file).replace('.xlsx', '_formatted.xlsx'))
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    ws.delete_rows(1)
    ws.delete_rows(1)
    ws.delete_rows(1)
    ## pure exception just for this table 
    ws.cell(1, column_index_from_string('B')).value = 'Туман (шаҳар) номи'
    ####
    for col in range(1, ws.max_column+1):
        for row in range(1, merge_walk_vert):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue 
            cell.alignment = Alignment(wrap_text=True,
                                       horizontal='center',
                                       vertical='center')
            cell.font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            cell.border = thin_border

            ######################## MERGE HEADERS #############################################
            horiz_merge = 0
            vert_merge = 0
            for walk_step in range(1, merge_walk_horiz-1): 
                col_next = col+walk_step      
                if col_next > ws.max_column:
                    break
                cell_next = ws.cell(row, col_next)
                if cell_next.value is None:
                    if row != 1:
                        try:
                            for upper_row in range(1, row):
                                
                                if not ws.cell(row-upper_row, col_next).value is None:
                                    raise Exception('break')
                            else:
                                horiz_merge += 1
                        except:
                            break

                           
                    else:
                        horiz_merge += 1
                else:
                    break

            for walk_step in range(1, merge_walk_vert-1): 
                row_next = row+walk_step      
                if row_next > merge_walk_vert-1:
                    break 
                cell_next = ws.cell(row_next, col)
                
                if cell_next.value is None:
                    vert_merge += 1
                else:
                    break
            
            # if ws.cell(row, col).value == '2024 йил' and row==4:
            #     print(ws.cell(row, col).value)
            #     print("Vert:", vert_merge)
            #     print("Horiz:", horiz_merge)
            #     start_column = get_column_letter(col)
            #     end_column = get_column_letter(col+horiz_merge)
            #     range_str = f'{start_column}{row}:{end_column}{row+vert_merge}'
            #     print(range_str)
            #     print()


            start_column = get_column_letter(col)
            end_column = get_column_letter(col+horiz_merge)
            range_str = f'{start_column}{row}:{end_column}{row+vert_merge}'
            ws.merge_cells(range_str)
         
###################################################################################

        ws.column_dimensions[get_column_letter(col)].width = 18.7
        ws.column_dimensions[get_column_letter(1)].width = 4.86
    
    for row in range(1, merge_walk_vert):   
        ws.row_dimensions[row].height = 37.25

    
    wb.save(filename_out)
    wb.close()











