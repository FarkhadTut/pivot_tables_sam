import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

def merge_headers(file, merge_walk_vert, sheet_name=None):
    merge_walk_horiz = 20

    wb = openpyxl.load_workbook(file)
    if sheet_name is not None:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    for col in range(1, ws.max_column+1):
        for row in range(1, merge_walk_vert):
            ######################## MERGE HEADERS #############################################
            horiz_merge = 0
            vert_merge = 0
            for walk_step in range(1, merge_walk_horiz-1): 
                col_next = col+walk_step      
                if col_next > ws.max_column:
                    break
                cell_next = ws.cell(row, col_next)
                if cell_next.value is None:
                    if row != 1:
                        try:
                            if ws.cell(row, col).value == '2024 йил' and row==4:
                                print(ws.cell(row-upper_row, col_next).value, row-upper_row, col_next)
                                exit()
                            for upper_row in range(1, row):
                                
                                if not ws.cell(row-upper_row, col_next).value is None:
                                    raise Exception('break')
                            else:
                                horiz_merge += 1
                        except:
                            break

                           
                    else:
                        horiz_merge += 1
                else:
                    break

            for walk_step in range(1, merge_walk_vert-1): 
                row_next = row+walk_step      
                if row_next > merge_walk_vert-1:
                    break 
                cell_next = ws.cell(row_next, col)
                
                if cell_next.value is None:
                    vert_merge += 1
                else:
                    break
            
            # if ws.cell(row, col).value == '2024 йил' and row==4:
            #     print(ws.cell(row, col).value)
            #     print("Vert:", vert_merge)
            #     print("Horiz:", horiz_merge)
            #     start_column = get_column_letter(col)
            #     end_column = get_column_letter(col+horiz_merge)
            #     range_str = f'{start_column}{row}:{end_column}{row+vert_merge}'
            #     print(range_str)
            #     print()


            start_column = get_column_letter(col)
            end_column = get_column_letter(col+horiz_merge)
            range_str = f'{start_column}{row}:{end_column}{row+vert_merge}'
            ws.merge_cells(range_str)
    wb.save(file)
    wb.close()            
###################################################################################